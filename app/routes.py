from flask import  render_template, redirect, flash, url_for, send_file, request
from app import app, db
from app.models import User, OvertimeReport, Profile
from app.auth import role_required, login_required, current_user
from app.forms import OvertimeReportForm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from sqlalchemy import desc





@app.route('/index')
@app.route('/')
def index():
    return render_template('index.html')








@app.route('/user_page', methods=['GET', 'POST'])
def user_page():
    return render_template('index.html')

@app.route('/overtime_report', methods=['GET', 'POST'])
def overtime_report():
    return render_template('overtime_report.html')





# маршуты для переработок
@app.route('/work_report', methods=['GET', 'POST'])
def work_report():
    return render_template('work_report.html')

# просмотр переработок пользователя с пагинацией
@app.route('/overwork/view_action', methods=['GET', 'POST'])
@login_required
def view_action():
    page = request.args.get('page', 1, type=int)  # Получаем номер страницы из запроса
    reports = OvertimeReport.query.filter_by(user_id=current_user.id)\
                                  .order_by(desc(OvertimeReport.task_date))\
                                  .paginate(page=page, per_page=10)  # Пагинация
    return render_template('overwork/view_action.html', reports=reports)


# удаление записей в бд для /overwork/view_action
@app.route('/overwork/delete_action/<int:report_id>', methods=['POST'])
@login_required
def delete_action(report_id):
    report = OvertimeReport.query.get_or_404(report_id)
    # Проверка, что текущий пользователь является владельцем записи
    if report.user_id != current_user.id:
        flash('Нет прав для удаления этой записи.', 'danger')
        return redirect(url_for('view_action'))
    db.session.delete(report)
    db.session.commit()
    flash('Запись успешно удалена.', 'success')
    return redirect(url_for('view_action'))

@app.route('/overwork/add_action', methods=['GET', 'POST'])
@login_required
def add_action():
    form = OvertimeReportForm()
    if form.validate_on_submit():
        report = OvertimeReport(
            project_name=form.project_name.data,
            task_description=form.task_description.data,
            # Преобразование даты из формата DD.MM.YYYY в YYYY-MM-DD
            task_date=datetime.strptime(form.task_date.data, '%d.%m.%Y').strftime('%Y-%m-%d'),
            day_type=form.day_type.data,
            hours_worked=form.hours_worked.data,
            user_id=current_user.id  # Предполагается, что у вас есть аутентификация
        )
        db.session.add(report)
        db.session.commit()
        flash('Overtime report has been added.', 'success')
        return redirect(url_for('add_action'))  # Перенаправление на главную страницу или другую страницу

    return render_template('/overwork/add_action.html', title='Add Overtime Report', form=form)

# загрузить
@app.route('/overwork/upload_action', methods=['GET', 'POST'])
def upload_action():
    return render_template('overwork/upload_action.html')


def calculate_k(tpr, tn, tc):
    """
    tпр - суммарное количество часов (время переработок)
    tн - нормальное месячное время - количество рабочих часов по производственному календарю
    tк - ставка на которой сотрудник работает
    """
    if tn == 0:
        raise ValueError("Normal monthly working hours (tn) cannot be zero.")
    if tc < 0 or tc > 1:
        raise ValueError("The rate (tc) must be between 0 and 1.")

    k = 1 + (2 * tpr) / (tn * tc)
    return k

# выгрузить
@app.route('/overwork/unload_action', methods=['GET', 'POST'])
@login_required
@role_required('is_admin', 'is_manager')
def unload_action():
    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        number_hours_int = int(request.form.get('number_hours'))

        # Попытка преобразования строк в объекты datetime
        try:
            start_date = datetime.strptime(start_date_str, '%d.%m.%Y').date()
            end_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
        except ValueError:
            # В случае ошибки возвращаем пользователя на форму с сообщением об ошибке
            flash('Неверный формат даты. Пожалуйста, используйте формат дд.мм.гггг', 'error')
            return redirect(url_for('unload_action'))  # Вернуться к форме

        # Проверка корректности диапазона дат
        if start_date > end_date:
            flash('Начальная дата не может быть позже конечной даты.', 'error')
            return redirect(url_for('unload_action'))  # Вернуться к форме

        # Получение данных из БД
        # Извлечение записей в заданном диапазоне дат
        reports = db.session.query(
            OvertimeReport,
            Profile.name,
            Profile.working_day
        ).join(Profile, OvertimeReport.user_id == Profile.user_id)\
        .filter(OvertimeReport.task_date >= start_date, OvertimeReport.task_date <= end_date)\
        .order_by(Profile.name, OvertimeReport.task_date)\
        .all()

                

        
        # Создание рабочей книги и листа
        wb = Workbook()
        ws = wb.active
        ws.title = "Overtime Reports"
        
        # Заголовки для файла Excel
        columns = [
            '№',
            'ФИО',
            'Проект',
            'Задача, которую выполнял',
            'Дата',
            'Выходной или рабочий день',
            'Количество сверхурочных часов',
            'Коэффициент'
        ]

        # Добавление заголовков в лист
        idx = 1
        for col_name in columns:
            ws.cell(row=1, column=idx, value=col_name)
            idx += 1

        # Словарь для хранения общего количества сверхурочных часов для каждого ФИО
        total_hours_per_profile = {}  
        # Словарь для хранения ставки сотрудника
        profile_working_day = {}

        # Добавление данных в лист
        row_idx = 2
        for report, profile_name, working_day in reports:
            profile_working_day[profile_name] = working_day
            total_hours_per_profile.setdefault(profile_name, 0)
            total_hours_per_profile[profile_name] += report.hours_worked
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=profile_name)  
            ws.cell(row=row_idx, column=3, value=report.project_name)
            ws.cell(row=row_idx, column=4, value=report.task_description)
            ws.cell(row=row_idx, column=5, value=report.task_date.strftime('%d.%m.%Y'))
            ws.cell(row=row_idx, column=6, value=report.day_type)
            ws.cell(row=row_idx, column=7, value=float(report.hours_worked))
            row_idx += 1
        

        # Добавление данных в лист и отслеживание интервалов для ФИО
        fio_ranges = {}
        row_idx = 2
        prev_profile_name = None
        for idx, (report, profile_name, working_day) in enumerate(reports, start=1):
            if profile_name != prev_profile_name:
                if prev_profile_name is not None:
                    fio_ranges[prev_profile_name]['end'] = row_idx - 1
                fio_ranges[profile_name] = {'start': row_idx}
                prev_profile_name = profile_name
            
            ws.cell(row=row_idx, column=1, value=idx)
            ws.cell(row=row_idx, column=2, value=profile_name)
            row_idx += 1
        
        # Закрытие последнего диапазона
        if prev_profile_name is not None:
            fio_ranges[prev_profile_name]['end'] = row_idx - 1

        # Объединение ячеек для каждого диапазона ФИО в столбце 'Коэффициент'
        coef_column = len(columns)  # Номер столбца 'Коэффициент' (последний столбец)
        # При добавлении данных в лист:
        for fio, rng in fio_ranges.items():
            start = rng['start']
            end = rng['end']
            tpr = total_hours_per_profile[fio]
                        # Предполагается, что profile_working_day хранит ставку сотрудника
            tc = profile_working_day[fio]  
            tn = number_hours_int
            k_value = calculate_k(tpr, tn, tc)

            if start != end:
                ws.merge_cells(start_row=start, start_column=coef_column, end_row=end, end_column=coef_column)
            ws.cell(row=start, column=coef_column, value=k_value)


        # Сохранение файла Excel в памяти
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        
        # Отправка файла пользователю
        virtual_workbook.seek(0)
        return send_file(virtual_workbook, as_attachment=True, download_name='overtime_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # Если это GET запрос, просто отображаем форму
    return render_template('overwork/unload_action.html')


@app.route('/admin_panel', methods=['GET', 'POST'])
@login_required  
@role_required('is_admin')
def admin_panel():
    return render_template('admin_panel.html')